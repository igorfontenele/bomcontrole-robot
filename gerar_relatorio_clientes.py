"""
gerar_relatorio_clientes.py
===========================
Gera um Excel com a base de clientes e faturamento médio mensal.

Fonte de dados:
  - Financeiro/PesquisaDetalhada  → movimentações por período (reusa caches
    do baixar_faturas.py: .cache_movimentacoes_*.json)
  - Cliente/Obter/{id}            → documento, contatos (email, telefone)
    Cacheado em .cache_clientes.json

Saída (Excel com 3 abas):
  - "Clientes"      : base completa ordenada por média mensal
  - "Top Clientes"  : ranking dos N principais (default 30)
  - "Resumo"        : totalizadores

Uso:
    pip install requests openpyxl
    python gerar_relatorio_clientes.py --ano 2026
    python gerar_relatorio_clientes.py --inicio 2025-01-01 --termino 2026-04-30
    python gerar_relatorio_clientes.py --ano 2026 --top 50 --saida clientes_2026.xlsx
"""

import argparse
import calendar
import json
import logging
import re
import time
from collections import defaultdict
from pathlib import Path

import requests

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Instale as dependências: pip install openpyxl")

# ---------------------------------------------------------------------------
# CONFIGURAÇÃO
# ---------------------------------------------------------------------------

import os
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

API_KEY  = os.getenv("BOMCONTROLE_API_KEY", "")
BASE_URL = os.getenv("BOMCONTROLE_BASE_URL",
                     "https://apinewintegracao.bomcontrole.com.br/integracao")

CACHE_DIR      = Path(".")
CACHE_CLIENTES = CACHE_DIR / ".cache_clientes.json"

# Parâmetros iguais aos do baixar_faturas.py — assim os caches de movimentações
# ficam compatíveis e podem ser reutilizados entre os dois scripts.
TIPO_DATA        = "DataCompetencia"
FILTRO_STATUS    = None
FILTRO_DESPESA   = False
ITENS_POR_PAGINA = 100

DELAY_LISTAGEM = 2.0
DELAY_CLIENTE  = 0.4
MAX_RETRIES    = 6

TOP_N_DEFAULT = 30

# Placeholders de "cliente não real" — ignorados no agrupamento
PLACEHOLDERS_CLIENTE = {
    "NÃO IDENTIFICADO", "NAO IDENTIFICADO", "", "cliente_desconhecido",
    "BRADESCO", "CAIXA ECONÔMICA FEDERAL", "CAIXA ECONOMICA FEDERAL",
    "ITAÚ UNIBANCO", "ITAU UNIBANCO", "SANTANDER",
    "SICOOB", "SICREDI", "BANCO DO BRASIL", "NUBANK",
}

# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Cache helpers
# ---------------------------------------------------------------------------

def _cache_load(path: Path):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            log.warning("Cache inválido em %s: %s", path, e)
    return {}


def _cache_save(path: Path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# HTTP com retry
# ---------------------------------------------------------------------------

def _request_com_retry(method, url, **kwargs):
    headers = kwargs.setdefault("headers", {})
    headers.setdefault("Authorization", f"ApiKey {API_KEY}")
    for tentativa in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.request(method, url, timeout=60, **kwargs)
            if resp.status_code == 429:
                retry_after = resp.headers.get("Retry-After")
                espera = float(retry_after) if retry_after else 5 * (2 ** (tentativa - 1))
                log.warning("429 — aguardando %.0fs (tentativa %d/%d)",
                            espera, tentativa, MAX_RETRIES)
                time.sleep(espera)
                continue
            if resp.status_code >= 500:
                espera = 5 * (2 ** (tentativa - 1))
                log.warning("HTTP %d — backoff %.0fs (tentativa %d/%d)",
                            resp.status_code, espera, tentativa, MAX_RETRIES)
                time.sleep(espera)
                continue
            resp.raise_for_status()
            return resp
        except requests.exceptions.ConnectionError as e:
            espera = 5 * (2 ** (tentativa - 1))
            log.warning("Conexão falhou (%s) — backoff %.0fs", e, espera)
            time.sleep(espera)
    raise RuntimeError(f"Falhou após {MAX_RETRIES} tentativas: {url}")


def _get(path, params=None):
    url = f"{BASE_URL}/{path.lstrip('/')}"
    return _request_com_retry("GET", url, params=params or {}).json()


# ---------------------------------------------------------------------------
# Listagem paginada — compatível com o cache do baixar_faturas.py
# ---------------------------------------------------------------------------

def listar_movimentacoes(ini, fim):
    cache_key  = f"{ini}|{fim}|{TIPO_DATA}|{FILTRO_STATUS}|{FILTRO_DESPESA}"
    cache_slug = re.sub(r'[^a-zA-Z0-9_-]', "_", f"{ini[:10]}_{fim[:10]}")
    cache_path = CACHE_DIR / f".cache_movimentacoes_{cache_slug}.json"

    cache = _cache_load(cache_path)
    if isinstance(cache, dict) and cache.get("cache_key") == cache_key:
        itens = cache.get("itens", [])
        log.info("Cache carregado: %d itens (%s → %s)", len(itens), ini[:10], fim[:10])
        return itens

    usa_detalhada = (FILTRO_STATUS is not None) or (FILTRO_DESPESA is not None)
    endpoint = "Financeiro/PesquisaDetalhada" if usa_detalhada else "Financeiro/Pesquisar"

    params_base = {
        "dataInicio":               ini,
        "dataTermino":              fim,
        "tipoData":                 TIPO_DATA,
        "paginacao.itensPorPagina": ITENS_POR_PAGINA,
    }
    if usa_detalhada:
        if FILTRO_STATUS  is not None: params_base["status"]  = FILTRO_STATUS
        if FILTRO_DESPESA is not None: params_base["despesa"] = str(FILTRO_DESPESA).lower()

    todas, pagina = [], 1
    while True:
        log.info("Buscando movimentações — página %d …", pagina)
        time.sleep(DELAY_LISTAGEM)
        data  = _get(endpoint, {**params_base, "paginacao.numeroDaPagina": pagina})
        itens = data.get("Itens", [])
        if not itens:
            break
        todas.extend(itens)
        total = data.get("TotalItens", len(todas))
        log.info("  %d / %d carregados", len(todas), total)
        _cache_save(cache_path, {"cache_key": cache_key, "itens": todas})
        if len(todas) >= total:
            break
        pagina += 1

    return todas


def coletar_por_ano(ano):
    """Processa mês a mês para aproveitar os caches do baixar_faturas.py."""
    todas = []
    for mes in range(1, 13):
        ultimo = calendar.monthrange(ano, mes)[1]
        ini = f"{ano:04d}-{mes:02d}-01 00:00:00"
        fim = f"{ano:04d}-{mes:02d}-{ultimo:02d} 23:59:59"
        log.info("──── %04d-%02d ────", ano, mes)
        todas.extend(listar_movimentacoes(ini, fim))
    return todas


# ---------------------------------------------------------------------------
# Cliente/Obter
# ---------------------------------------------------------------------------

def obter_cliente(id_cliente, cache):
    k = str(id_cliente)
    entry = cache.get(k)
    if isinstance(entry, dict) and not entry.get("erro"):
        return entry
    try:
        time.sleep(DELAY_CLIENTE)
        data = _get(f"Cliente/Obter/{id_cliente}")
        cache[k] = data
        _cache_save(CACHE_CLIENTES, cache)
        return data
    except Exception as e:
        log.warning("  Cliente/%s → %s", id_cliente, e)
        cache[k] = {"erro": True, "msg": str(e)[:200]}
        _cache_save(CACHE_CLIENTES, cache)
        return None


# ---------------------------------------------------------------------------
# Formatação
# ---------------------------------------------------------------------------

def _formatar_doc(doc):
    if not doc:
        return ""
    d = re.sub(r'\D', "", str(doc))
    if len(d) == 11:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return d or str(doc)


def _formatar_tel(tel):
    if not tel:
        return ""
    d = re.sub(r'\D', "", str(tel))
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:]}"
    if len(d) == 10:
        return f"({d[:2]}) {d[2:6]}-{d[6:]}"
    return str(tel)


def _mes_key(s):
    return str(s)[:7] if s else None


# ---------------------------------------------------------------------------
# Agregação
# ---------------------------------------------------------------------------

def agregar_por_cliente(movs):
    """Agrupa por IdCliente somando valor e contando meses distintos com receita."""
    agg = defaultdict(lambda: {
        "id_cliente": None,
        "nome_lista": "",
        "total":      0.0,
        "meses":      set(),
        "qtd_mov":    0,
    })
    ignorados = 0
    for m in movs:
        idc = m.get("IdCliente")
        if not idc or int(idc) <= 0:
            ignorados += 1
            continue
        nome = (m.get("NomeFantasiaClienteFornecedor")
                or m.get("NomeClienteFornecedor")
                or "").strip()
        if nome.upper() in PLACEHOLDERS_CLIENTE:
            ignorados += 1
            continue
        a = agg[idc]
        a["id_cliente"] = idc
        a["nome_lista"] = nome
        mk = _mes_key(m.get("DataCompetencia")
                      or m.get("DataVencimento")
                      or m.get("DataPadrao"))
        if mk:
            a["meses"].add(mk)
        try:
            a["total"] += float(m.get("Valor") or 0)
        except Exception:
            pass
        a["qtd_mov"] += 1
    log.info("Movimentações agregadas: %d | ignoradas (sem cliente real): %d",
             sum(x["qtd_mov"] for x in agg.values()), ignorados)
    return agg


def enriquecer_com_contatos(agg, cache_clientes):
    linhas = []
    total = len(agg)
    for i, (idc, info) in enumerate(sorted(agg.items(), key=lambda kv: -kv[1]["total"]), 1):
        log.info("[%d/%d] Cliente %s — %s", i, total, idc, info["nome_lista"][:50])
        cli = obter_cliente(idc, cache_clientes)

        nome = info["nome_lista"]
        tipo, doc_raw = "", ""
        emails, telefones = [], []

        if cli and not cli.get("erro"):
            tipo = cli.get("TipoPessoa") or ""
            pj = cli.get("PessoaJuridica") or None
            pf = cli.get("PessoaFisica") or None
            if pj:
                nome    = pj.get("NomeFantasia") or pj.get("RazaoSocial") or nome
                doc_raw = pj.get("Documento") or ""
            elif pf:
                nome    = pf.get("Nome") or nome
                doc_raw = pf.get("Documento") or ""
            for c in (cli.get("Contatos") or []):
                e = (c.get("Email") or "").strip()
                t = (c.get("Telefone") or "").strip()
                if e and e not in emails:
                    emails.append(e)
                if t and t not in telefones:
                    telefones.append(t)

        qtd_meses = len(info["meses"]) or 1
        media     = info["total"] / qtd_meses

        linhas.append({
            "id":              idc,
            "nome":            nome,
            "tipo":            tipo,
            "documento":       _formatar_doc(doc_raw),
            "emails":          "; ".join(emails),
            "telefones":       "; ".join(_formatar_tel(t) for t in telefones),
            "total_faturado":  round(info["total"], 2),
            "meses_ativos":    qtd_meses,
            "media_mensal":    round(media, 2),
            "qtd_movimentacoes": info["qtd_mov"],
        })
    return linhas


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

_FILL_HEADER = PatternFill("solid", fgColor="305496")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_FMT_MONEY = '"R$" #,##0.00'


def _estilo_header(ws, row=1):
    for cell in ws[row]:
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_HEADER


def _largura_colunas(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def escrever_excel(linhas, top_n, saida: Path, periodo_label: str):
    ordenado = sorted(linhas, key=lambda x: x["media_mensal"], reverse=True)
    wb = Workbook()

    # ───── Aba 1: Clientes ─────
    ws = wb.active
    ws.title = "Clientes"
    ws.append([
        "Id", "Nome", "Tipo", "CPF/CNPJ", "Emails", "Telefones",
        "Total Faturado", "Meses Ativos", "Média Mensal", "Qtd Mov.",
    ])
    _estilo_header(ws)
    for l in ordenado:
        ws.append([
            l["id"], l["nome"], l["tipo"], l["documento"],
            l["emails"], l["telefones"],
            l["total_faturado"], l["meses_ativos"],
            l["media_mensal"], l["qtd_movimentacoes"],
        ])
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for c in row:
            c.number_format = _FMT_MONEY
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for c in row:
            c.number_format = _FMT_MONEY
    _largura_colunas(ws, [8, 42, 10, 22, 42, 24, 16, 12, 16, 10])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ───── Aba 2: Top Clientes ─────
    ws2 = wb.create_sheet("Top Clientes")
    ws2.append([
        "#", "Nome", "CPF/CNPJ", "Emails", "Telefones",
        "Média Mensal", "Total Faturado", "Meses Ativos",
    ])
    _estilo_header(ws2)
    for i, l in enumerate(ordenado[:top_n], 1):
        ws2.append([
            i, l["nome"], l["documento"], l["emails"], l["telefones"],
            l["media_mensal"], l["total_faturado"], l["meses_ativos"],
        ])
    for row in ws2.iter_rows(min_row=2, min_col=6, max_col=7):
        for c in row:
            c.number_format = _FMT_MONEY
    _largura_colunas(ws2, [5, 42, 22, 42, 24, 16, 16, 12])
    ws2.freeze_panes = "A2"

    # ───── Aba 3: Resumo ─────
    ws3 = wb.create_sheet("Resumo")
    total_geral   = sum(l["total_faturado"] for l in linhas)
    meses_max     = max((l["meses_ativos"] for l in linhas), default=0)
    media_geral   = total_geral / meses_max if meses_max else 0
    clientes_ativos = len(linhas)

    ws3["A1"] = "Resumo — Base de Clientes BomControle"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.merge_cells("A1:B1")

    linhas_resumo = [
        ("Período analisado",              periodo_label),
        ("Clientes únicos com faturamento", clientes_ativos),
        ("Total faturado no período",       total_geral),
        ("Meses cobertos (máx.)",           meses_max),
        ("Média de faturamento mensal",     media_geral),
    ]
    for i, (k, v) in enumerate(linhas_resumo, start=3):
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=v)

    ws3["B5"].number_format = _FMT_MONEY   # total faturado
    ws3["B7"].number_format = _FMT_MONEY   # média mensal
    _largura_colunas(ws3, [36, 28])

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    log.info("✔ Arquivo gerado: %s", saida.resolve())


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args():
    p = argparse.ArgumentParser(
        description="Gera Excel com a base de clientes e faturamento médio mensal."
    )
    g = p.add_mutually_exclusive_group()
    g.add_argument("--ano", metavar="AAAA",
                   help="Ano completo (processa 12 meses reusando caches).")
    g.add_argument("--inicio", metavar="AAAA-MM-DD",
                   help="Início de intervalo livre (use com --termino)")
    p.add_argument("--termino", metavar="AAAA-MM-DD",
                   help="Fim de intervalo livre (obrigatório com --inicio)")
    p.add_argument("--top", type=int, default=TOP_N_DEFAULT,
                   help=f"Quantidade de clientes no ranking (default {TOP_N_DEFAULT})")
    p.add_argument("--saida", default="relatorio_clientes.xlsx",
                   help="Arquivo de saída (default relatorio_clientes.xlsx)")
    return p.parse_args()


def main():
    args = _parse_args()

    if args.ano:
        try:
            ano = int(args.ano)
        except ValueError:
            raise SystemExit("Formato inválido para --ano. Use AAAA, ex: 2026")
        log.info("Período: ano completo de %d", ano)
        movs = coletar_por_ano(ano)
        periodo = f"{ano}-01-01 → {ano}-12-31"
    elif args.inicio:
        if not args.termino:
            raise SystemExit("--inicio requer --termino.")
        ini = f"{args.inicio} 00:00:00"
        fim = f"{args.termino} 23:59:59"
        log.info("Período: %s → %s", args.inicio, args.termino)
        movs = listar_movimentacoes(ini, fim)
        periodo = f"{args.inicio} → {args.termino}"
    else:
        raise SystemExit("Passe --ano AAAA  OU  --inicio AAAA-MM-DD --termino AAAA-MM-DD")

    log.info("Total de movimentações coletadas: %d", len(movs))

    log.info("Agregando por cliente...")
    agg = agregar_por_cliente(movs)
    log.info("Clientes únicos com receitas: %d", len(agg))

    if not agg:
        raise SystemExit("Nenhum cliente encontrado no período informado.")

    cache_clientes = _cache_load(CACHE_CLIENTES)
    linhas = enriquecer_com_contatos(agg, cache_clientes)

    escrever_excel(linhas, args.top, Path(args.saida), periodo)


if __name__ == "__main__":
    main()
